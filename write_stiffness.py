#  -*- coding: gbk -*-

from openpyxl import load_workbook

#写入桩刚度
def write_stiffness(num,stiffness):

    wb_1 = load_workbook('paramenter.xlsx')
    ws_1 = wb_1.active

    for j in ws_1.iter_rows(min_row=2):
        if j[0].value < num:
            continue
        elif j[0].value == num:
            j[9].value = stiffness
            wb_1.save('paramenter.xlsx')
        else:
            break

#测试
#write_stiffness(197,2000)