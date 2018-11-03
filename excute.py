# -*- coding: gbk -*-

from stiffness import stiffness
from write_stiffness import write_stiffness
from openpyxl import load_workbook


wb_1 = load_workbook('paramenter.xlsx')
ws_1 = wb_1.active

for row in ws_1.iter_rows(min_row=2):
    num = row[0].value
    stiff = stiffness(num)
    write_stiffness(num,stiff)

print('done!')


# a = stiffness(300)
# write_stiffness(300,a)
# b = stiffness(195)
# write_stiffness(195,b)
# print(a,b)