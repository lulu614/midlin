from openpyxl import load_workbook

def creat_soil(num_begin,num_last,soil_layer):

    wb_1 = load_workbook('soil.xlsx')
    ws_1 = wb_1.active

    count = ws_1.max_row +1
    for num in range(num_begin,num_last+1):
        for layer in soil_layer:
            ws_1.cell(row=count, column=1).value = num
            ws_1.cell(row=count, column=2).value = layer[0]
            ws_1.cell(row=count, column=3).value = layer[1]
            ws_1.cell(row=count, column=4).value = layer[2]
            count += 1

    wb_1.save('soil.xlsx')

#excute
creat_soil(194,314,((2,30000,0.7),(7,40000,0.6),(10,30000,0.5),(8,35000,0.7)))