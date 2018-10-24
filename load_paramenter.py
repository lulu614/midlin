import math
from openpyxl import load_workbook


#导入全部参数
def load_paramenter():

    #打开Excel表格
    wb_1 = load_workbook('paramenter.xlsx')
    ws_1 = wb_1.active
    wb_2 = load_workbook('soil.xlsx')
    ws_2 = wb_2.active

    pile_dict = {}
    for i in ws_1.iter_rows(min_row=2):
        num = i[0].value
        x = i[1].value/1000
        y = i[2].value/1000
        a = i[3].value
        b = i[4].value
        Q = i[5].value
        l = i[6].value
        d = i[7].value

        # 导入所在位置所有土层厚度&Es压缩模量&v泊松比
        soil_layer = []
        for j in ws_2.iter_rows(min_row=2):
            if j[0].value < num:
                continue
            elif j[0].value == num:
                h = j[1].value
                Es = j[2].value
                v = j[3].value
                soil_layer.append((h,Es,v))         # 土层厚度&Es压缩模量&v泊松比
            else:
                break
        pile_dict[num] = [x,y,a,b,Q,l,d,soil_layer]
    return pile_dict



#找到范围内的桩坐标
def sereach_num(num,pile_dict):
    num_list = []
    x_0 = pile_dict[num][0]
    y_0 = pile_dict[num][1]
    for i in pile_dict.keys():
        x_1 = pile_dict[i][0]
        y_1 = pile_dict[i][1]
        r = math.sqrt((x_0 - x_1) ** 2 + (y_0 - y_1) ** 2)
        if r < 5 and r > 0:
            num_list.append([i,r])
    return num_list



#获得对应编号的范围内桩的所有参数
#num  桩编号
#pile_dict  全部参数字典
def search_num(num,pile_dict):
    pile_dict_choose = {}
    temp = pile_dict[num][2:]
    temp.append(0)
    pile_dict_choose[num] = temp

    num_list = sereach_num(num,pile_dict)
    for i in num_list:
        n = i[0]
        r = i[1]
        temp = pile_dict[n][2:]
        temp.append(r)
        pile_dict_choose[n] = temp

#pile_dict_choose  [a,b,Q,l,d,soil_layer,r]
    return pile_dict_choose


#测试
# pile_dict_choose = search_num(249,load_paramenter())
# print(pile_dict_choose)